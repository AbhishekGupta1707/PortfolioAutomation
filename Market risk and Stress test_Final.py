#!/usr/bin/env python
# coding: utf-8

# In[1]:


import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import yfinance as yf
from scipy.stats import norm
import os
from openpyxl import load_workbook,Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill
from email.message import EmailMessage
import smtplib
import time
import gc


#pip install --upgrade yfinance


# In[2]:


nse_list = pd.read_csv('https://archives.nseindia.com/content/equities/EQUITY_L.csv')
nse_list = nse_list.apply(lambda col: col.map(lambda x: x.strip() if isinstance(x, str) else x))
nse_list.columns = nse_list.columns.str.strip()

nse_list = nse_list[nse_list['SERIES'] == 'EQ']
nse_list = nse_list.drop_duplicates(subset='SYMBOL')

if 'LAST_PRICE' in nse_list.columns:
    nse_list['LAST_PRICE'] = pd.to_numeric(nse_list['LAST_PRICE'], errors='coerce')
    nse_list = nse_list[nse_list['LAST_PRICE'] > 50]

if 'TOTTRDQTY' in nse_list.columns:
    nse_list['TOTTRDQTY'] = pd.to_numeric(nse_list['TOTTRDQTY'], errors='coerce')
    nse_list = nse_list[nse_list['TOTTRDQTY'] > 100000]

stocks = nse_list['SYMBOL'].tolist()[:300]

print(f"Stocks after static filter: {len(stocks)}")
print(f"Total Stocks after static filter: {len(nse_list)}")

#Data download
def chunk_list(lst, size):
    for i in range(0, len(lst), size):
        yield lst[i:i + size]

all_price_data = []
for batch in chunk_list(stocks, 20):
    tickers = [x + '.NS' for x in batch]
    try:
        df = yf.download(
            tickers,
            start='2015-01-01',
            group_by='ticker',
            threads=False,
            progress=False
        )
        all_price_data.append(df)
    except Exception as e:
        print(f"Batch error: {e}")
    
    time.sleep(2)
    gc.collect()

data = pd.concat(all_price_data, axis=1)
print(f"Validate total Stocks after static filter: {len(data)}")
print("Download complete")

#Factor calculations
results = []
for stock in stocks:
    try:
        df = data[stock + '.NS']['Close'].dropna()
        if len(df) < 200:
            continue
        returns = df.pct_change().dropna()
        momentum = (df.iloc[-1] / df.iloc[-126]) - 1
        volatility = returns.std() * (252 ** 0.5)
        cum = (1 + returns).cumprod()
        drawdown = (cum / cum.cummax() - 1).min()
        results.append([stock, momentum, volatility, drawdown])
    except:
        continue
df_scores = pd.DataFrame(results, columns=['Stock', 'Momentum', 'Volatility', 'Drawdown'])
print(f"Stocks after data filter: {len(df_scores)}")

#Dynamic filters
df_scores = df_scores[df_scores['Momentum'] > -0.1]
vol_threshold = df_scores['Volatility'].quantile(0.8)
df_scores = df_scores[df_scores['Volatility'] < vol_threshold]

print(f"Stocks after momentum + vol filter: {len(df_scores)}")

#Final score
def normalize(x):
    return (x - x.min()) / (x.max() - x.min() + 1e-9)

df_scores['mom_n'] = normalize(df_scores['Momentum'])
df_scores['vol_n'] = normalize(df_scores['Volatility'])
df_scores['Drawdown_abs'] = df_scores['Drawdown'].abs()
df_scores['dd_n'] = normalize(df_scores['Drawdown_abs'])
df_scores['sharpe'] = df_scores['Momentum'] / (df_scores['Volatility'] + 1e-9)
df_scores['sharpe_n'] = normalize(df_scores['sharpe'])

df_scores['BUY_SCORE'] = (
    0.4 * df_scores['mom_n'] +
    0.3 * df_scores['sharpe_n'] +
    0.3 * (1 - df_scores['dd_n'])
)
df_scores = df_scores.sort_values('BUY_SCORE', ascending=False)

#Enrichment
final_rows = []

total_capital = 1000000      # 10L
max_pct = 0.05               # 5% per stock

for _, row in df_scores.iterrows():
    stock = row['Stock']
    try:
        ticker = yf.Ticker(stock + '.NS')
        info = ticker.info
        pe = info.get('trailingPE', np.nan)
        if pd.notna(pe) and pe > 40:
            val_penalty = 0.1
        else:
            val_penalty = 0
        adj_score = row['BUY_SCORE'] - val_penalty
        
        df = data[stock + '.NS']['Close'].dropna()
        price = df.iloc[-1]
        low_52w = df.tail(252).min()
        
        market_cap = info.get('marketCap', np.nan)
        if pd.isna(market_cap) or market_cap < 10000 * 1e7:  # ₹20,000 Cr
            continue
        sector = info.get('sector', 'Unknown')
        
        dist_52w = (price / low_52w) - 1
        
        #Reason engine
        if row['Momentum'] > 0.3 and row['Volatility'] < 0.3:
            reason = "Strong momentum + Low risk"
        elif row['Momentum'] > 0.2:
            reason = "Momentum play"
        elif dist_52w < 0.2:
            reason = "Near 52W low recovery"
        elif row['Volatility'] > 0.5:
            reason = "High risk"
        else:
            reason = "Stable compounder"
        
        #Position sizing
        max_alloc = total_capital * max_pct
        risk_factor = 1 / (row['Volatility'] + 1e-9)
        adj_alloc = max_alloc * (risk_factor / 10)   # scale down
        max_units = int(adj_alloc / price)
        
        final_rows.append([
            stock, sector, market_cap,
            price, low_52w, dist_52w,
            row['Momentum'], row['Volatility'], row['Drawdown'],
            adj_score, reason,
            max_alloc, max_units
        ])
    except:
        continue

df_final = pd.DataFrame(final_rows, columns=[
    'INSTRUMENT','SECTOR','MARKET_CAP',
    'PRICE','52W_LOW','DIST_FROM_52W_LOW',
    'MOMENTUM','VOLATILITY','DRAWDOWN',
    'BUY_SCORE','REASON',
    'MAX_ALLOC','MAX_UNITS'
])

#Output
df_final = df_final.sort_values('BUY_SCORE', ascending=False)

df_final = df_final[[
    'INSTRUMENT','SECTOR','MARKET_CAP',
    'PRICE','52W_LOW','DIST_FROM_52W_LOW',
    'MOMENTUM','VOLATILITY','DRAWDOWN',
    'BUY_SCORE','REASON',
    'MAX_ALLOC','MAX_UNITS'
]]

df_final=df_final[['INSTRUMENT','SECTOR','MARKET_CAP','PRICE','52W_LOW','BUY_SCORE','REASON','DIST_FROM_52W_LOW','MOMENTUM','VOLATILITY','DRAWDOWN']]

df_final['MARKET_CAP'] = pd.to_numeric(df_final['MARKET_CAP'])
df_final['MARKET_CAP'] = (df_final['MARKET_CAP'] / 10000000).round(2)
df_final['PRICE'] = pd.to_numeric(df_final['PRICE']).round(2)
df_final['52W_LOW'] = pd.to_numeric(df_final['52W_LOW']).round(2)
df_final['BUY_SCORE'] = pd.to_numeric(df_final['BUY_SCORE']).round(2)
df_final['DIST_FROM_52W_LOW'] = df_final['PRICE'] - df_final['52W_LOW']
df_final['DIST_FROM_52W_LOW'] = pd.to_numeric(df_final['DIST_FROM_52W_LOW']).round(2)
df_final['MCAP/PRICE'] = df_final['MARKET_CAP'] / df_final['PRICE']
df_final['MCAP/PRICE'] = pd.to_numeric(df_final['MCAP/PRICE']).round(2)
df_final['MOMENTUM'] = pd.to_numeric(df_final['MOMENTUM']).round(2)
df_final['VOLATILITY'] = pd.to_numeric(df_final['VOLATILITY']).round(2)
df_final['DRAWDOWN'] = pd.to_numeric(df_final['DRAWDOWN']).round(2)

df_final.columns = ['INSTRUMENT','SECTOR','MARKET_CAP[Cr.]','PRICE','52W_LOW','BUY_SCORE','REASON','DIST_FROM_52W_LOW','MCAP/PRICE','MOMENTUM','VOLATILITY','DRAWDOWN']

df_final.head(30).to_excel("Top_Buy_Stocks.xlsx", index=False)

print("Final enriched stock list saved to Top_Buy_Stocks.xlsx")


# In[3]:


holdings=pd.read_excel('portfolio_holdings.xlsx',dtype='str')
holdings = holdings.apply(lambda col: col.map(lambda x: x.strip() if isinstance(x, str) else x))

temp=df_final[['INSTRUMENT','PRICE','52W_LOW']]
temp.columns=['INSTRUMENT','LTP','52W LOW']
master=pd.concat([holdings,temp])
master['NET POS']=master['NET POS'].fillna(1)
master['AVG COST']=master['AVG COST'].fillna(master['LTP'])
master['CHG %']=master['CHG %'].fillna(0)
master['TOTAL P&L']=master['TOTAL P&L'].fillna(0)
master['TOTAL P&L%']=master['TOTAL P&L%'].fillna(0)
master['INV AMT']=master['INV AMT'].fillna(master['LTP'])
master['MKT VAL']=master['MKT VAL'].fillna(master['LTP'])
master['WEIGHTED AVG RATE']=master['WEIGHTED AVG RATE'].fillna(master['LTP'])
master['P&L DAY']=master['P&L DAY'].fillna(0)
master['SOURCE'] =master['INSTRUMENT'].isin(holdings['INSTRUMENT'])
master['SOURCE']=master['SOURCE'].map({True: 'Existing', False: 'New'})

master['MKT VAL']=pd.to_numeric(master['MKT VAL'])


# In[4]:


cl_95=0.95
cl_99=0.99
tail_q=0.01
n=10
report = pd.DataFrame({
    "Metric": [
        "Start_date",
        f"Historical VaR {int(cl_95*100)}%",
        f"Parametric VaR {int(cl_95*100)}% (Normal)",
        f"Tail loss (worst {int(tail_q*100)}%) avg",
        "Worst Drawdown",
        "Stress PnL (-5%)",
        f"Worst {n} days avg PnL",
    ]
})

for x in master['INSTRUMENT']:
    #df=yf.download(x,start='2010-01-01')['Close']

    df = yf.download(x + '.NS', start='2010-01-01', progress=False)['Close']
    if df.empty:
        df = yf.download(x + '.BO', start='2010-01-01', progress=False)['Close']
    
    df=df.squeeze()
    returns=df.pct_change().dropna() #YTM
    #print(f"{df}'\n\n'{returns}")

    #getting start date
    df.index = pd.to_datetime(df.index)
    start_date=df.index[0]
    
    #Return diagnostics
    summary = pd.DataFrame({
        'mean_daily': [returns.mean()],
        'vol_daily': [returns.std()],
        'min_daily': [returns.min()],
        'max_daily': [returns.max()],
        'n_obs': [returns.shape[0]],
    })
    #print(summary)
    #print('')
    
    plt.figure()
    returns.hist(bins=100)
    plt.title('Daily Returns Histogram')
    plt.xlabel('Ddaily returns')
    plt.ylabel('Frequency')
    #plt.savefig(f"Graphs/Daily Returns Histogram_{x}.png")
    plt.close()
    
    #Historical VaR
    var_hist_95=np.percentile(returns, (1-cl_95) * 100)
    var_hist_99=np.percentile(returns, (1-cl_99) * 100)
    
    HistoricalVaR=pd.DataFrame({
        'VaR type' : ['Historical','Historical'],
        'Confidence' : [cl_95,cl_99],
        'VaR (return)' : [var_hist_95,var_hist_99],
    })
    #print(HistoricalVaR)
    #print('')
    
    #Parametric Var
    mu=returns.mean()
    sigma=returns.std()
    
    var_param_95 = norm.ppf(1-cl_95, loc=mu, scale=sigma)
    var_param_99 = norm.ppf(1-cl_99, loc=mu, scale=sigma)
    
    ParametricVaR=pd.DataFrame({
        'VaR type' : ['Parametric (Normal)','Parametric (Normal)'],
        'Confidence' : [cl_95,cl_99],
        'VaR (return)' : [var_param_95,var_param_99],  
    })
    #print(ParametricVaR)
    #print('')
    
    compare_var=pd.DataFrame({
        'Historical VaR (return)' : [var_hist_95,var_hist_99],
        'Parametric VaR (return)' : [var_param_95,var_param_99],
    },index=[f"{int(cl_95*100)}%", f"{int(cl_99*100)}%"])
    #print(compare_var)
    #print('')
    #Normal distribution assign low probability to extreme tail events. As a result, this leads Parametric Var to be lower compared to emperical Var at 99% CI
    #Parametric VaR underestimates risk at very high confidence internval due to fat tails
    #Historical VaR hence captures losses that normal dist. VaR smoothers away in extreme loss events
    
    #VaR Backtesting
    var_threshold=var_hist_95
    
    exceptions=returns<var_threshold
    exception_rate=exceptions.mean()
    
    var_backtesting=pd.DataFrame({
        'VaR model' : ['Historical 95% (constant)'],
        'VaR threshold (return)' : [var_threshold],
        'exception_rate' : [exception_rate],
        'exception_rate' : [1-cl_95],
        'n_exceptions' : [exceptions.sum()],
        'n_total' : [len(exceptions)],
    })
    #print(var_backtesting)
    #print('')
    
    #Tail risk
    tail_cutoff=returns.quantile(tail_q)
    tail_returns=returns[returns<=tail_cutoff]
    
    tail_stats=pd.DataFrame({
        'tail_quantile' : [tail_q],
        'tail_cutoff (return)' : [tail_cutoff],
        'avg_tail_returns' : [tail_returns.mean()],
        'min_tail_returns' : [tail_returns.min()],
        'n_tail_obs' : [len(tail_returns)],
    })
    #print(tail_stats)
    #print('')
    
    plt.figure()
    tail_returns.hist(bins=30)
    plt.title('Tail Returns histogram (Worst 1%)')
    plt.xlabel('Return')
    plt.ylabel('Frequency')
    #plt.savefig(f"Graphs/Tail Returns histogram (Worst 1%)_{x}.png")
    plt.close()
    
    #Stress testing
    portfolio_value=master.loc[master['INSTRUMENT'] == x, 'MKT VAL'].values[0]
    
    shocks=[-0.03,-0.05,-0.10]
    
    stress_table=pd.DataFrame({
        'shock_returns' : shocks,
        'stress_pnl' : [shock*portfolio_value for shock in shocks],
    })
    #print(stress_table)
    #print('')
    
    #hist_stress
    worst_days=returns.nsmallest(n)
    hist_stress=pd.DataFrame({
        'data' : worst_days.index,
        'return' : worst_days.values,
        'pnl' : worst_days.values*portfolio_value,
    }).reset_index(drop=True)
    #print(hist_stress)
    #print('')
    
    #Drawdowns
    cum=(1+returns).cumprod()
    dd=cum/cum.cummax()-1
    
    dd_stats=pd.DataFrame({
        'max_drawdown' : [dd.min()],
        'max_drawdown_pn' : [dd.min()*portfolio_value]
    })
    #print(dd_stats)
    
    plt.figure()
    dd.plot()
    plt.title('Drawdown')
    plt.xlabel('Date')
    plt.ylabel('Drawdown')
    #plt.savefig(f"Graphs/Drawdown_{x}.png")
    plt.close()
    
    #Risk Report
    report[f"{x.split('.')[0]}"]=[
        start_date,
        var_hist_95*portfolio_value,
        var_param_95*portfolio_value,
        tail_returns.mean()*portfolio_value,
        dd.min()*portfolio_value,
        -0.05*portfolio_value,
        hist_stress["pnl"].mean() if len(hist_stress) > 0 else np.nan
    ]
    #print(report)
print('Report generated.')

reportoutput=report.T
reportoutput.columns=reportoutput.iloc[0]
reportoutput=reportoutput.iloc[1:]

reportoutput.columns=['StartDate','HistVaR95%','ParamVaR95%(Normal)','TailLoss(worst1%)Avg','WorstDrawdown','StressPnL(-5%)','Worst10DaysAvgPnL']
reportoutput = reportoutput.reset_index()
reportoutput.rename(columns={'index': 'Script'}, inplace=True)

reportoutput['StartDate'] = pd.to_datetime(reportoutput['StartDate']).dt.date
reportoutput['HistVaR95%'] = pd.to_numeric(reportoutput['HistVaR95%']).round(2)
reportoutput['ParamVaR95%(Normal)'] = pd.to_numeric(reportoutput['ParamVaR95%(Normal)']).round(2)
reportoutput['TailLoss(worst1%)Avg'] = pd.to_numeric(reportoutput['TailLoss(worst1%)Avg']).round(2)
reportoutput['WorstDrawdown'] = pd.to_numeric(reportoutput['WorstDrawdown']).round(2)
reportoutput['StressPnL(-5%)'] = pd.to_numeric(reportoutput['StressPnL(-5%)']).round(2)
reportoutput['Worst10DaysAvgPnL'] = pd.to_numeric(reportoutput['Worst10DaysAvgPnL']).round(2)

master['NET POS'] = pd.to_numeric(master['NET POS']).round(2)
master['INV AMT'] = pd.to_numeric(master['INV AMT']).round(2)
master['MKT VAL'] = pd.to_numeric(master['MKT VAL']).round(2)
master['AVG COST'] = pd.to_numeric(master['AVG COST']).round(2)
master['52W LOW'] = pd.to_numeric(master['52W LOW']).round(2)
master['LTP'] = pd.to_numeric(master['LTP']).round(2)
master['CHG %'] = pd.to_numeric(master['CHG %']).round(2)

reportoutput=pd.merge(reportoutput,master[['INSTRUMENT','NET POS','INV AMT','MKT VAL','AVG COST','52W LOW','LTP','CHG %','SOURCE']],how='left',left_on='Script',right_on='INSTRUMENT')
reportoutput['PotentialWorst'] = abs(reportoutput['WorstDrawdown'])
reportoutput['PnL/Worst'] = ((reportoutput['MKT VAL'] - reportoutput['INV AMT'])/(abs(reportoutput['WorstDrawdown']) * reportoutput['MKT VAL'] + 1e-9))

reportoutput=reportoutput[['INSTRUMENT','SOURCE','NET POS','INV AMT','MKT VAL','AVG COST','52W LOW','LTP','PotentialWorst','PnL/Worst','StartDate','HistVaR95%','ParamVaR95%(Normal)','TailLoss(worst1%)Avg','WorstDrawdown','StressPnL(-5%)','Worst10DaysAvgPnL','CHG %' ]]

# Load data
input_file = reportoutput
output_file = "PortfolioSummary.xlsx"

data = input_file

def normalize(series):
    return (series - series.min()) / (series.max() - series.min() + 1e-9)

data['norm_tail'] = data['TailLoss(worst1%)Avg'].rank(pct=True)
data['norm_drawdown'] = data['WorstDrawdown'].rank(pct=True)
data['norm_far'] = data['PotentialWorst'].rank(pct=True)

def get_driver(row):    
    # Priority-based clean reasoning (ONLY TOP DRIVER)
    if row['PnL/Worst'] < 1 and row['norm_far'] > 0.6:
        return "Weak + More downside left"
    elif row['norm_tail'] > 0.7:
        return "High crash risk"
    elif row['norm_drawdown'] > 0.7:
        return "Deep drawdown history"
    elif row['CHG %'] < 0:
        return "Negative momentum"
    elif row['MKT VAL']-row['INV AMT'] > 0:
        return "Profitable (hold)"
    elif row['PotentialWorst'] < 50:
        return "Near bottom (limited downside)"
    else:
        return "Stable"
data['Driver'] = data.apply(get_driver, axis=1)

def sell_score(row):
    score = 0
    reasons = []
    # 1. Risk Efficiency
    if row['PnL/Worst'] < 1:
        score += 2
        reasons.append("(PnL/Worst <1)")
    # 2. Tail Risk
    tail_score = 1.5 * row['norm_tail']
    score += tail_score
    if tail_score > 1:
        reasons.append("High tail risk")
    # 3. Drawdown
    dd_score = 1.0 * row['norm_drawdown']
    score += dd_score
    if dd_score > 0.7:
        reasons.append("High drawdown risk")
    # 4. Future Risk
    far_score = 2.0 * row['norm_far']
    score += far_score
    if far_score > 1.5:
        reasons.append("High downside left")
    # 5. Momentum
    if row['CHG %'] < 0:
        score += 1
        reasons.append("Negative momentum")
    # 6. Position Size
    if row['MKT VAL'] > data['MKT VAL'].median():
        score += 1
        reasons.append("Large position size")

    if row['MKT VAL']-row['INV AMT'] > 0:
        score -= 3
        reasons.append("Winner (protected)")
    if row['PotentialWorst'] < 50:
        score -= 2
        reasons.append("Near bottom (protected)")
    if row['LTP'] <= row['52W LOW'] * 1.1:
        score -= 2
        reasons.append("Near 52W low")
    return round(score, 2), ", ".join(reasons)

data[['Score', 'Score Drivers']] = data.apply(lambda row: pd.Series(sell_score(row)), axis=1)

def classify_reason(row):
    if row['Score'] >= 6:
        return "Clean"
    elif row['norm_tail'] < 0.3 and row['norm_drawdown'] < 0.3:
        return "Core"
    elif row['norm_tail'] < 0.6:
        return "Growth"
    else:
        return "Tactical"

data['Reason'] = data.apply(classify_reason, axis=1)

def action(row):
    # Strong exit
    if row['Score'] >= 6:
        return "SELL"
    # Weak, exit on bounce
    elif row['Score'] >= 4:
        return "SELL ON BOUNCE"
    # NEW stock → potential BUY (with quality filter)
    elif (
        row['SOURCE'] == 'New' and 
        row['Score'] < 4 and 
        row['LTP'] <= row['52W LOW'] * 1.25
    ):
        return "BUY"
    # Existing winner → add more
    elif row['Score'] < 2 and (row['MKT VAL'] - row['INV AMT']) > 0:
        return "ADD"
    # Default
    else:
        return "HOLD"

data['Action'] = data.apply(action, axis=1)

def buy_score(row):
    score = 0
    # Efficiency
    if row['PnL/Worst'] > 1:
        score += 2
    # Low risk preferred
    score += (1 - row['norm_tail']) * 2
    score += (1 - row['norm_drawdown']) * 1.5
    # Momentum
    if row['CHG %'] > 0:
        score += 1
    # Avoid broken stocks
    if row['LTP'] > row['52W LOW'] * 1.2:
        score += 1
    return round(score, 2)
data.head()
data.to_excel("AnalyticalData.xlsx", index=False)

print('Data file writing completed.')

data=data[['INSTRUMENT','SOURCE','NET POS','INV AMT','MKT VAL','Driver','Score','Score Drivers','Reason','Action','AVG COST','52W LOW','PotentialWorst','PnL/Worst','TailLoss(worst1%)Avg','WorstDrawdown','StressPnL(-5%)','Worst10DaysAvgPnL' ]]

data['NET POS']=pd.to_numeric(data['NET POS']).copy()
data['INV AMT']=pd.to_numeric(data['INV AMT'])
data['MKT VAL']=pd.to_numeric(data['MKT VAL'])
data['Score']=pd.to_numeric(data['Score'])
data['AVG COST']=pd.to_numeric(data['AVG COST'])
data['52W LOW']=pd.to_numeric(data['52W LOW'])

data['PotentialWorst']=pd.to_numeric(data['PotentialWorst'])
data['PnL/Worst']=pd.to_numeric(data['PnL/Worst'])
data['TailLoss(worst1%)Avg']=pd.to_numeric(data['TailLoss(worst1%)Avg'])
data['WorstDrawdown']=pd.to_numeric(data['WorstDrawdown'])
data['StressPnL(-5%)']=pd.to_numeric(data['StressPnL(-5%)'])
data['Worst10DaysAvgPnL']=pd.to_numeric(data['Worst10DaysAvgPnL'])

#Output
new=data[data['SOURCE']=='New']
new = new[new['Action'].isin(['BUY','ADD','HOLD'])]
existing=data[data['SOURCE']=='Existing']
data=pd.concat([new,existing])

order = ['BUY','ADD','HOLD','SELL ON BOUNCE','SELL']
data['Action'] = pd.Categorical(data['Action'], categories=order, ordered=True)
data = data.sort_values(by=['SOURCE','Action'], ascending=[False, False])

#data.to_excel(output_file, sheet_name='SUMMARY', index=False)
#print(f" Summary created: {output_file}")


# In[5]:


def classify(row):
    score_drivers = str(row.get('Score Drivers',''))
    if row['Score'] > 6:
        return 'SELL'
    elif row['Score'] > 4:
        return 'SELL ON RISE'
    elif row['SOURCE'] == 'New' and row['Score'] < 3.5 and row['Driver'] != 'High crash risk':
        return 'BUY'
    elif row['SOURCE'] == 'Existing' and row['Score'] < 2 and 'Winner (protected)' in score_drivers:
        return 'ADD'
    else:
        return 'HOLD'
data['Final_Action'] = data.apply(classify, axis=1)

#Excel Section-wise Output
wb = Workbook()
ws = wb.active
ws.title = "SUMMARY"
sections = ['BUY', 'ADD', 'HOLD', 'SELL ON RISE', 'SELL']
columns = data.columns
row_cursor = 1

section_colors = {
    'BUY': '92D050',          # Green
    'ADD': '00B0F0',          # Blue
    'HOLD': 'FFFF00',         # Yellow
    'SELL ON RISE': 'FFC000', # Orange
    'SELL': 'FF0000'          # Red
}

for section in sections:
    df_section = data[data['Final_Action'] == section].copy()
    if df_section.empty:
        continue
    df_section = df_section.reindex(columns=columns)
    # Clean data
    df_section = df_section.replace([float('inf'), -float('inf')], None)
    df_section = df_section.astype(object).fillna('')
    # Section header in Excel
    cell = ws.cell(row=row_cursor, column=1, value=section)
    cell.font = Font(bold=True, color="000000")
    cell.fill = PatternFill(start_color=section_colors[section], end_color=section_colors[section], fill_type="solid")
    row_cursor += 1
    # Write table
    for r in dataframe_to_rows(df_section, index=False, header=True):
        ws.append(r)
        row_cursor += 1
    row_cursor += 2

wb.save(output_file)

#HTML Email Section-wise Output
email_columns = ['INSTRUMENT','SOURCE','Driver','Score Drivers','Reason','Final_Action',
                 'NET POS','INV AMT','MKT VAL','AVG COST','52W LOW',
                 'PotentialWorst','PnL/Worst','WorstDrawdown']

html_content = ""
for section in sections:
    df_section = data[data['Final_Action'] == section]
    if df_section.empty:
        continue
    html_content += f"""
    <h2>{section}</h2>
    {df_section[email_columns].to_html(index=False, border=1)}
    <br><br>
    """

# Wrap full HTML with CSS
html_content = f"""
<html>
<head>
<style>
body {{
    font-family: Arial, sans-serif;
}}
table {{
    border-collapse: collapse;
    width: 100%;
    font-size: 12px;
}}
th {{
    background-color: #f2f2f2;
    padding: 8px;
    border: 1px solid #ddd;
}}
td {{
    padding: 6px;
    text-align: center;
    border: 1px solid #ddd;
}}
h2 {{
    color: #2E86C1;
    border-bottom: 2px solid #2E86C1;
    padding-bottom: 4px;
}}
</style>
</head>
<body>
<h1 style="text-align:center;">Portfolio Summary</h1>
{html_content}
</body>
</html>
"""


# In[6]:


'''
#Buy script email:
# Convert dataframe to HTML
html_table = df_final.to_html(index=False)

sender_email = "abhishek1994g@gmail.com"
receiver_email = "abhishek1994g@gmail.com"
app_password = "uqqa poqs vuct evup"  # NOT your Gmail password

msg = EmailMessage()
msg['Subject'] = "Portfolio Buy Summary"
msg['From'] = sender_email
msg['To'] = receiver_email

msg.add_alternative(f"""
<html>
<body>
<h3>Portfolio Summary</h3>
{html_table}
</body>
</html>
""", subtype='html')

with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
    smtp.login(sender_email, app_password)
    smtp.send_message(msg)

print(" Email sent with table")
'''
#Sell script email:
# Convert dataframe to HTML
html_table = data[['INSTRUMENT','SOURCE','Driver','Score Drivers','Reason','Action','NET POS','INV AMT','MKT VAL','AVG COST','52W LOW','PotentialWorst','PnL/Worst','WorstDrawdown']].to_html(index=False)
sender_email = "abhishek1994g@gmail.com"
receiver_email = "abhishek1994g@gmail.com"
app_password = os.getenv("EMAIL_PASS")

msg = EmailMessage()
msg['Subject'] = "Portfolio Summary"
msg['From'] = sender_email
msg['To'] = receiver_email

'''
msg.add_alternative(f"""
<html>
<body>
<h3>Portfolio Summary</h3>
{html_table}
</body>
</html>
""", subtype='html')
'''
msg.add_alternative(html_content, subtype='html')

with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
    smtp.login(sender_email, app_password)
    smtp.send_message(msg)

print(" Email sent with table")

